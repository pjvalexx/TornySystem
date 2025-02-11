import openpyxl
from openpyxl.styles import Font
from conexion.conexionBD import connectionBD
import os

def generar_reporte_inventario():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM materials"
                cursor.execute(querySQL)
                materiales = cursor.fetchall()

        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Escribir encabezados
        headers = ["ID", "Nombre", "Descripción", "Cantidad", "Unidad", "Proveedor", "Stock Mínimo"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Escribir datos
        for row_num, material in enumerate(materiales, 2):
            ws.cell(row=row_num, column=1, value=material['id'])
            ws.cell(row=row_num, column=2, value=material['name'])
            ws.cell(row=row_num, column=3, value=material['description'])
            ws.cell(row=row_num, column=4, value=material['quantity'])
            ws.cell(row=row_num, column=5, value=material['unit'])
            ws.cell(row=row_num, column=6, value=material['supplier_id'])
            ws.cell(row=row_num, column=7, value=material['minimum_stock'])

        # Guardar el archivo en una ubicación válida
        file_path = os.path.join(os.getcwd(), "reporte_inventario.xlsx")
        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"Error en generar_reporte_inventario: {e}")
        return None

def generar_reporte_ordenes():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM work_orders"
                cursor.execute(querySQL)
                ordenes = cursor.fetchall()

        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Órdenes de Trabajo"

        # Escribir encabezados
        headers = ["ID", "Cliente ID", "Tipo de Servicio", "Descripción", "Fecha de Entrega", "Monto", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Escribir datos
        for row_num, orden in enumerate(ordenes, 2):
            ws.cell(row=row_num, column=1, value=orden['id'])
            ws.cell(row=row_num, column=2, value=orden['client_id'])
            ws.cell(row=row_num, column=3, value=orden['service_type'])
            ws.cell(row=row_num, column=4, value=orden['description'])
            ws.cell(row=row_num, column=5, value=orden['delivery_date'])
            ws.cell(row=row_num, column=6, value=orden['amount'])
            ws.cell(row=row_num, column=7, value=orden['status'])

        # Guardar el archivo en una ubicación válida
        file_path = os.path.join(os.getcwd(), "reporte_ordenes.xlsx")
        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"Error en generar_reporte_ordenes: {e}")
        return None
    
def generar_copia_seguridad():
    try:
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()

        # Obtener los datos de los clientes
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM clients"
                cursor.execute(querySQL)
                clientes = cursor.fetchall()

        # Crear una hoja para los clientes
        ws_clientes = wb.create_sheet(title="Clientes")
        headers_clientes = ["ID", "Nombre", "Dirección", "Teléfono", "Documento"]
        for col_num, header in enumerate(headers_clientes, 1):
            cell = ws_clientes.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        for row_num, cliente in enumerate(clientes, 2):
            ws_clientes.cell(row=row_num, column=1, value=cliente['id'])
            ws_clientes.cell(row=row_num, column=2, value=cliente['name'])
            ws_clientes.cell(row=row_num, column=3, value=cliente['address'])
            ws_clientes.cell(row=row_num, column=4, value=cliente['phone'])
            ws_clientes.cell(row=row_num, column=5, value=cliente['documento'])

        # Obtener los datos de las órdenes de trabajo
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM work_orders"
                cursor.execute(querySQL)
                ordenes = cursor.fetchall()

        # Crear una hoja para las órdenes de trabajo
        ws_ordenes = wb.create_sheet(title="Órdenes de Trabajo")
        headers_ordenes = ["ID", "Cliente ID", "Tipo de Servicio", "Descripción", "Fecha de Entrega", "Monto", "Estado"]
        for col_num, header in enumerate(headers_ordenes, 1):
            cell = ws_ordenes.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        for row_num, orden in enumerate(ordenes, 2):
            ws_ordenes.cell(row=row_num, column=1, value=orden['id'])
            ws_ordenes.cell(row=row_num, column=2, value=orden['client_id'])
            ws_ordenes.cell(row=row_num, column=3, value=orden['service_type'])
            ws_ordenes.cell(row=row_num, column=4, value=orden['description'])
            ws_ordenes.cell(row=row_num, column=5, value=orden['delivery_date'])
            ws_ordenes.cell(row=row_num, column=6, value=orden['amount'])
            ws_ordenes.cell(row=row_num, column=7, value=orden['status'])

        # Obtener los datos de los proveedores
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM suppliers"
                cursor.execute(querySQL)
                proveedores = cursor.fetchall()

        # Crear una hoja para los proveedores
        ws_proveedores = wb.create_sheet(title="Proveedores")
        headers_proveedores = ["ID", "Nombre", "Dirección", "Teléfono", "Email"]
        for col_num, header in enumerate(headers_proveedores, 1):
            cell = ws_proveedores.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        for row_num, proveedor in enumerate(proveedores, 2):
            ws_proveedores.cell(row=row_num, column=1, value=proveedor['id'])
            ws_proveedores.cell(row=row_num, column=2, value=proveedor['name'])
            ws_proveedores.cell(row=row_num, column=3, value=proveedor['address'])
            ws_proveedores.cell(row=row_num, column=4, value=proveedor['phone'])
            ws_proveedores.cell(row=row_num, column=5, value=proveedor['email'])

        # Obtener los datos de los materiales
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT * FROM materials"
                cursor.execute(querySQL)
                materiales = cursor.fetchall()

        # Crear una hoja para los materiales
        ws_materiales = wb.create_sheet(title="Materiales")
        headers_materiales = ["ID", "Nombre", "Descripción", "Cantidad", "Unidad", "Stock Mínimo", "Proveedor ID"]
        for col_num, header in enumerate(headers_materiales, 1):
            cell = ws_materiales.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        for row_num, material in enumerate(materiales, 2):
            ws_materiales.cell(row=row_num, column=1, value=material['id'])
            ws_materiales.cell(row=row_num, column=2, value=material['name'])
            ws_materiales.cell(row=row_num, column=3, value=material['description'])
            ws_materiales.cell(row=row_num, column=4, value=material['quantity'])
            ws_materiales.cell(row=row_num, column=5, value=material['unit'])
            ws_materiales.cell(row=row_num, column=6, value=material['minimum_stock'])
            ws_materiales.cell(row=row_num, column=7, value=material['supplier_id'])

        # Guardar el archivo en una ubicación válida
        file_path = os.path.join(os.getcwd(), "copia_seguridad.xlsx")
        wb.save(file_path)
        return file_path
    except Exception as e:
        print(f"Error en generar_copia_seguridad: {e}")
        return None